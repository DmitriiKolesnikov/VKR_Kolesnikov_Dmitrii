import pandas as pd
from openpyxl import load_workbook

wb = load_workbook('Copy of ДАТА_МОДЕЛИ.xlsx', data_only=True)
ws = wb['Лог регрессия']

year_per_col = {}
quarter_per_col = {}
current_year = None
for col in range(3, 43):
    cell_year = ws.cell(row=10, column=col).value
    if cell_year is not None:
        current_year = cell_year
    year_per_col[col] = current_year

    cell_quarter = ws.cell(row=11, column=col).value
    if isinstance(cell_quarter, str):
        if cell_quarter.startswith('I '):
            quarter_per_col[col] = 1
        elif cell_quarter.startswith('II '):
            quarter_per_col[col] = 2
        elif cell_quarter.startswith('III '):
            quarter_per_col[col] = 3
        elif cell_quarter.startswith('IV '):
            quarter_per_col[col] = 4
        else:
            quarter_per_col[col] = None
    else:
        quarter_per_col[col] = cell_quarter

bank_names = []
for row in range(16, ws.max_row + 1):
    name = ws.cell(row=row, column=1).value
    if name and name not in bank_names:
        bank_names.append(name)

records = []
for bank in bank_names:
    rows_for_bank = [
        r for r in range(16, ws.max_row + 1)
        if ws.cell(row=r, column=1).value == bank
    ]
    var_row = { ws.cell(row=r, column=2).value: r for r in rows_for_bank }

    default_in_year = {}
    default_row = var_row.get('Дефолт')
    if default_row is not None:
        for col in range(3, 43):
            year = year_per_col[col]
            if year is None or year < 2002 or year > 2011:
                continue
            val = ws.cell(row=default_row, column=col).value
            if val is True or val == 'True':
                default_in_year[year] = True
    for col in range(3, 43):
        year = year_per_col[col]
        if year is None or year < 2002 or year > 2011:
            continue
        default_in_year.setdefault(year, False)

    for col in range(3, 43):
        year = year_per_col[col]
        quarter = quarter_per_col[col]
        if year is None or quarter is None or year < 2003 or year > 2011:
            continue

        inflation    = ws.cell(row=13, column=col).value
        gdp_growth   = ws.cell(row=14, column=col).value
        key_rate     = ws.cell(row=15, column=col).value
        H1           = ws.cell(row=var_row['Норматив достаточности капитала (Н1)'], column=col).value
        npl          = ws.cell(row=var_row['Доля просроченных кредитов (NPL%)'], column=col).value
        roa          = ws.cell(row=var_row['Рентабельность активов (ROA%)'], column=col).value
        h3           = ws.cell(row=var_row['Коэффициент ликвидности (Н3%)'], column=col).value
        size         = ws.cell(row=var_row['Размер банка (лог активов)'], column=col).value
        share        = ws.cell(row=var_row['Доля банка в активах системы'], column=col).value
        loan_growth  = ws.cell(row=var_row['Темп роста кредитного портфеля (YoY)'], column=col).value
        LDR          = ws.cell(row=var_row['Loan-to-Deposit Ratio (LDR, %)'], column=col).value
        interbank    = ws.cell(row=var_row['Доля межбанковских заимствований в пассивах (%)'], column=col).value
        profit_growth= ws.cell(row=var_row['Прирост прибыли банка (YoY)'], column=col).value
        ownership    = ws.cell(row=var_row['Тип собственности банка'], column=col).value

        cycle_stage  = ws.cell(row=9, column=col).value
        default_flag = 1 if default_in_year.get(year+1, False) else 0

        if any(v == 'DEFAULT' for v in [
            H1, npl, roa, h3, size,
            share, loan_growth, LDR,
            interbank, profit_growth, ownership
        ]):
            continue

        records.append({
            "Название банка": bank,
            "Год": int(year),
            "Квартал": int(quarter),
            "Инфляция": inflation,
            "Темп прироста ВВП": gdp_growth,
            "Ключевая ставка": key_rate,
            "Норматив достаточности капитала (Н1)": H1,
            "Доля просроченных кредитов (NPL%)": npl,
            "Рентабельность активов (ROA%)": roa,
            "Коэффициент ликвидности (Н3%)": h3,
            "Размер банка (лог активов)": size,
            "Доля банка в активах системы": share,
            "Темп роста кредитного портфеля (YoY)": loan_growth,
            "Loan-to-Deposit Ratio (LDR, %)": LDR,
            "Доля межбанковских заимствований в пассивах (%)": interbank,
            "Прирост прибыли банка (YoY)": profit_growth,
            "Тип собственности банка": ownership,
            "Этап экономического цикла": cycle_stage,
            "Дефолт": default_flag
        })

df = pd.DataFrame(records)

df_full   = df
df_rise   = df[df["Этап экономического цикла"] == 'Подъем']
df_peak   = df[df["Этап экономического цикла"] == 'Пик']
df_fall   = df[df["Этап экономического цикла"] == 'Спад']
df_bottom = df[df["Этап экономического цикла"] == 'Дно']

for name, subset in [
    ('data_full.csv',   df_full),
    ('data_rise.csv',   df_rise),
    ('data_peak.csv',   df_peak),
    ('data_fall.csv',   df_fall),
    ('data_bottom.csv', df_bottom),
]:
    subset.to_csv(name, encoding='utf-8-sig', index=False)
